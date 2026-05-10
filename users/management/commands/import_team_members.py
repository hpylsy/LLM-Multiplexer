import random
import re
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from pypinyin import Style, lazy_pinyin

from api_keys.models import APIKey
from users.models import Profile


PINYIN_MAP = {
    "黄海": "HuangHai",
    "薄汉博": "BoHanbo",
    "黄沛玥": "HuangPeiyue",
}


def normalize_group(value):
    value = str(value or "").strip()
    mapping = {
        "电控": Profile.GROUP_CONTROL,
        "算法": Profile.GROUP_ALGO,
        "机械": Profile.GROUP_MECH,
        "宣传": Profile.GROUP_MEDIA,
    }
    if value not in mapping:
        raise ValueError(f"未知分组: {value}")
    return mapping[value]


def normalize_member_type(value):
    value = str(value or "").strip()
    if value not in {Profile.MEMBER_CORE, Profile.MEMBER_RESERVE}:
        raise ValueError(f"未知队员身份: {value}")
    return value


def build_account_name(name):
    if name in PINYIN_MAP:
        return PINYIN_MAP[name]
    sanitized = re.sub(r"\s+", "", str(name or "").strip())
    if re.fullmatch(r"[A-Za-z][A-Za-z0-9]*", sanitized):
        return sanitized
    pinyin_parts = lazy_pinyin(sanitized, style=Style.NORMAL)
    if not pinyin_parts:
        raise ValueError(f"姓名 {name} 无法转换为拼音")
    return "".join(part.capitalize() for part in pinyin_parts)


def generate_cliproxy_key(account_name):
    chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    prefix = "".join(random.choice(chars) for _ in range(7))
    return f"sk-{prefix}{account_name}"


class Command(BaseCommand):
    help = "从 Excel 导入队员名单，批量创建门户账号并绑定 CLIProxy API Key"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Excel 文件路径")
        parser.add_argument("--sheet", type=str, default=None, help="工作表名称，默认读取第一个")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"文件不存在: {file_path}")

        workbook = load_workbook(file_path)
        sheet = workbook[options["sheet"]] if options["sheet"] else workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Excel 文件为空")

        header = [str(item).strip() if item is not None else "" for item in rows[0]]
        required = ["姓名", "组别", "队员身份"]
        missing = [item for item in required if item not in header]
        if missing:
            raise CommandError(f"缺少必要列: {', '.join(missing)}")

        created_count = 0
        skipped_count = 0

        for raw in rows[1:]:
            row = dict(zip(header, raw))
            if not any(row.values()):
                continue
            name = str(row.get("姓名") or "").strip()
            if not name:
                continue

            account_name = build_account_name(name)
            group = normalize_group(row.get("组别"))
            member_type = normalize_member_type(row.get("队员身份"))
            grade = str(row.get("年级") or "").strip()
            if not grade:
                match = re.search(r"20\d{2}", name)
                grade = match.group(0) if match else ""

            if User.objects.filter(username=account_name).exists():
                skipped_count += 1
                continue

            with transaction.atomic():
                user = User.objects.create_user(username=account_name, password=account_name)
                profile = user.profile
                profile.display_name = name
                profile.lab_group = group
                profile.member_type = member_type
                profile.grade = grade
                profile.save()

                raw_key = generate_cliproxy_key(account_name)
                metadata = APIKey.build_token_metadata(raw_key)
                APIKey.objects.create(
                    user=user,
                    name="API 密钥",
                    token_hash=metadata["token_hash"],
                    token_prefix=metadata["token_prefix"],
                    token_masked=metadata["token_masked"],
                    token_plaintext=raw_key,
                    source=APIKey.SOURCE_CLIPROXY,
                    status=APIKey.STATUS_ACTIVE,
                    note="批量导入创建",
                )
                created_count += 1

        self.stdout.write(self.style.SUCCESS(f"批量导入完成：创建 {created_count} 个，跳过 {skipped_count} 个"))
